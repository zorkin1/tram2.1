import csv
import numpy
import openpyxl
import pandas

#функция поиска максимума строк и колонок
def max_row_colum ():

    book = openpyxl.open("C://tram/справочник.xlsx")
    sheet = book
    sheet = book.active

    rows = sheet.max_row
    cols = sheet.max_column

    print ('максимум строк', rows, 'максимум колонок', cols)

def read_csv():
    with open('C://tram/mgt_example_trans.csv') as file:
        keys = file.readline().strip().split(',')
        return [dict(zip(keys, line.strip().split(','))) for line in file]

from os import path
import openpyxl
from openpyxl import load_workbook
#def reading_files():
#    full_name = path.basename(r'C:\tram\mgt_example_trans.csv ')
#    name = path.splitext(full_name)
#    print("прочитан файл", name)
#   f = open('C://tram/mgt_example_trans.csv ', 'r')
#   print(f.read())
#   f.close()
#book_scv = openpyexcel.open("mgt_example_trans.csv", read_only=True)


book = openpyxl.open("C://tram/справочник.xlsx")
sheet = book.active
#print({sheet.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True)}
#col = 1
#for i in range(1, 40):
#    print(f' значение строке файла {i} колонки {col} равно значение  {sheet.cell(row=i,column=col).value}')

print( {sheet.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True)})
#           {sheet.cell(row=i,column=col).value},
##         {sheet.cell(row=i,column=col+2).value},
#             {sheet.cell(row=i,column=col+3).value},
#           {sheet.cell(row=i,column=col+4).value},
#           {sheet.cell(row=i,column=col+5).value},
#           {sheet.cell(row=i,column=col+6).value},
#           {sheet.cell(row=i,column=col+7).value},
#           {sheet.cell(row=i,column=col+8).value},
#           {sheet.cell(row=i,column=col+9).value},
#           {sheet.cell(row=i,column=col+10).value},
#           {sheet.cell(row=i,column=col+10).value}
#)









